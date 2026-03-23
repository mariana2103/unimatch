#!/usr/bin/env python3
"""
DGES Course Data Builder
========================
1. Loads the authoritative 2026/27 course list from dados_dges/ (official DGES files).
2. Fetches each course's detail page from www.dges.gov.pt for provas, pesos,
   notas mínimas, district, and historical cutoffs (cached in ./cache/).
3. Merges everything and writes dges_cursos_completo.xlsx:
     Sheet "Cursos"           — one row per course, all fields
     Sheet "Provas (detalhe)" — one row per exam requirement
     Sheet "Histórico"        — one row per (course, year, fase) grade entry

Sources of truth (in priority order for each field):
  - Nota último colocado 2025 (1ª fase) → iesip_vagas_2026-2027_nota_ultimo_colocado_*
  - Nota último colocado 2024 (1ª fase) → dges_vagascna_nota_ult_colocado_*
  - Historical notas 2023-2025           → DGES detail page stats table
  - Provas / pesos / notas mínimas       → DGES detail page
  - Vagas 2026/27                        → iesip_vagas_2026-2027_pares_*
  - Tipo (pública/privada)               → pares file SUBSISTEMA column
  - Área CNAEF                           → pares file COD CNAEF column

Usage:
    pip install -r requirements.txt
    python scrape_dges.py

Re-run freely — cached HTML pages are not re-fetched.
Delete ./cache/ to force a full refresh.
"""

import io
import logging
import re
import time
from pathlib import Path

import openpyxl
import requests
from bs4 import BeautifulSoup
from openpyxl.styles import Alignment, Font, PatternFill

# ── Config ────────────────────────────────────────────────────────────────────

SCRIPT_DIR = Path(__file__).parent
ROOT_DIR   = SCRIPT_DIR.parent
DATA_DIR   = ROOT_DIR / "dados_dges"
CACHE_DIR  = SCRIPT_DIR / "cache"
OUTPUT     = SCRIPT_DIR / "dges_cursos_completo.xlsx"

DELAY     = 0.4
MAX_RETRY = 3

PARES_FILE    = DATA_DIR / "iesip_vagas_2026-2027_pares_ies_cursos_16.02.2026v2_.xlsx"
NOTA_FILE     = DATA_DIR / "iesip_vagas_2026-2027_nota_ultimo_colocado_1afase_2025_16.02.2026_.xlsx"
NOTA2024_FILE = DATA_DIR / "dges_vagascna_nota_ult_colocado_1afase2024_2025_17.02.2025.xlsx"

DETAIL_URL = "https://www.dges.gov.pt/guias/detcursopi.asp?codc={codc}&code={code}"

CACHE_DIR.mkdir(exist_ok=True)

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s  %(levelname)-8s %(message)s",
    datefmt="%H:%M:%S",
)
log = logging.getLogger(__name__)

SESSION = requests.Session()
SESSION.headers.update({"User-Agent": "DGES-research/2.0 (educational)"})

# ── CNAEF → app area ──────────────────────────────────────────────────────────

AREA_MAP = [
    ((480, 489), "Informática e Dados"),
    ((500, 539), "Engenharia e Tecnologia"),
    ((540, 599), "Engenharia e Tecnologia"),
    ((720, 729), "Ciências da Vida e Saúde"),
    ((640, 649), "Ciências da Vida e Saúde"),
    ((420, 429), "Ciências da Vida e Saúde"),
    ((430, 469), "Ciências Exatas e da Natureza"),
    ((340, 349), "Economia, Gestão e Contabilidade"),
    ((380, 389), "Direito, Ciências Sociais e Humanas"),
    ((310, 319), "Direito, Ciências Sociais e Humanas"),
    ((220, 229), "Direito, Ciências Sociais e Humanas"),
    ((100, 109), "Educação e Desporto"),
    ((810, 819), "Educação e Desporto"),
    ((210, 219), "Artes e Design"),
    ((200, 299), "Artes e Design"),
]

def cnaef_to_area(code) -> str:
    try:
        c = int(code)
    except (ValueError, TypeError):
        return ""
    for (lo, hi), area in AREA_MAP:
        if lo <= c <= hi:
            return area
    return ""

# ── Value helpers ─────────────────────────────────────────────────────────────

def _s(v) -> str:
    s = str(v).strip() if v is not None else ""
    return "" if s in ("None", "nan") else s

def _f(v) -> float | None:
    if v is None or v == "":
        return None
    try:
        import math
        f = float(v)
        return None if math.isnan(f) else f
    except (TypeError, ValueError):
        return None

def _i(v) -> int | None:
    f = _f(v)
    return None if f is None else int(f)

def _to_float(v) -> float | None:
    if v is None:
        return None
    try:
        s = str(v).strip().replace(",", ".")
        f = float(re.sub(r"[^\d.]", "", s))
        return f if f > 0 else None
    except Exception:
        return None

# ── HTTP fetch with cache ─────────────────────────────────────────────────────

def _cache_key(url: str) -> str:
    return re.sub(r"[^\w]", "_", url)[:180]

def fetch_cached(url: str) -> bytes:
    path = CACHE_DIR / _cache_key(url)
    if path.exists():
        return path.read_bytes()
    for attempt in range(MAX_RETRY):
        try:
            r = SESSION.get(url, timeout=30)
            r.raise_for_status()
            path.write_bytes(r.content)
            time.sleep(DELAY)
            return r.content
        except Exception as exc:
            if attempt == MAX_RETRY - 1:
                raise
            time.sleep(2 ** attempt)

# ── Excel loaders ─────────────────────────────────────────────────────────────

def _xlsx_rows(path: Path, sheet: str, header_row: int = 0) -> list[dict]:
    wb = openpyxl.load_workbook(str(path), data_only=True, read_only=True)
    ws = wb[sheet]
    rows = list(ws.iter_rows(values_only=True))
    wb.close()
    if len(rows) <= header_row:
        return []
    headers = rows[header_row]
    return [
        dict(zip(headers, r))
        for r in rows[header_row + 1:]
        if any(v is not None for v in r)
    ]

def load_pares() -> list[dict]:
    rows = _xlsx_rows(PARES_FILE, "PUB_PRIV26272", header_row=0)
    log.info("Pares: %d courses.", len(rows))
    return rows

def load_notas_2025() -> dict[tuple[str, str], float]:
    col = "NOTA ULTIMO COLOCADO REGIME GERAL DE ACESSO 2025/2026"
    rows = _xlsx_rows(NOTA_FILE, "Concurso Nacional", header_row=0)
    out: dict[tuple[str, str], float] = {}
    for r in rows:
        uo    = _s(r.get("COD UO"))
        curso = _s(r.get("COD CURSO"))
        nota  = _f(r.get(col))
        if uo and curso and nota is not None:
            out[(uo, curso)] = nota
    log.info("Notas 2025: %d entries.", len(out))
    return out

def load_notas_2024() -> dict[tuple[str, str], float]:
    col = "Nota último colocado 1ª Fase 2024 (cont. geral)"
    rows = _xlsx_rows(NOTA2024_FILE, "Nacional", header_row=3)
    out: dict[tuple[str, str], float] = {}
    for r in rows:
        inst  = _s(r.get("Código Instit."))
        curso = _s(r.get("Código Curso"))
        nota  = _f(r.get(col))
        if inst and curso and nota is not None:
            out[(inst, curso)] = nota
    log.info("Notas 2024: %d entries.", len(out))
    return out

# ── HTML parser ───────────────────────────────────────────────────────────────

def _sections(soup: BeautifulSoup) -> dict[str, list]:
    out: dict[str, list] = {}
    for h2 in soup.find_all("h2"):
        title = h2.get_text(strip=True)
        siblings = []
        for sib in h2.next_siblings:
            if getattr(sib, "name", None) == "h2":
                break
            siblings.append(sib)
        out[title] = siblings
    return out

def _sec_text(sections: dict, fragment: str) -> str:
    for k, v in sections.items():
        if fragment.lower() in k.lower():
            return " ".join(
                n.get_text(" ", strip=True) if hasattr(n, "get_text") else str(n).strip()
                for n in v
            ).strip()
    return ""

def _sec_nodes(sections: dict, fragment: str) -> list:
    for k, v in sections.items():
        if fragment.lower() in k.lower():
            return v
    return []

def scrape_detail(cod_curso: str, cod_uo: str) -> dict:
    url = DETAIL_URL.format(codc=cod_curso, code=cod_uo)
    try:
        raw = fetch_cached(url)
    except Exception as exc:
        log.debug("  Skip %s/%s — %s", cod_uo, cod_curso, exc)
        return {}

    try:
        soup = BeautifulSoup(raw, "lxml")
    except Exception:
        return {}

    sections = _sections(soup)
    detail: dict = {}

    # ── District ──────────────────────────────────────────────────────────────
    addr = _sec_text(sections, "Endere")
    m = re.search(
        r"\d{4}-\d{3}\s+([A-ZÁÉÍÓÚÂÊÔÃÕÀÇ][A-ZÁÉÍÓÚÂÊÔÃÕÀÇ a-záéíóúâêôãõàç]{2,40}?)"
        r"(?:\s*\bMap|\s*\bTel|\s*<|\s*$)",
        addr,
    )
    if m:
        detail["distrito"] = m.group(1).strip().title()

    # ── Tipo from características (HTML is ground truth for private schools) ──
    caract = _sec_text(sections, "Caracter")
    detail["tipo"] = "privada" if "Privado" in caract else "publica"

    # ── Provas de ingresso ────────────────────────────────────────────────────
    provas_nodes = _sec_nodes(sections, "Provas de Ingresso")
    provas: list[dict] = []
    if provas_nodes:
        frag = BeautifulSoup("".join(str(n) for n in provas_nodes), "lxml")
        for br in frag.find_all("br"):
            br.replace_with("\n")
        conjunto_id = 1
        for line in frag.get_text().splitlines():
            clean = line.replace("\xa0", " ").strip()
            if not clean or clean.lower().startswith("um dos"):
                continue
            if re.fullmatch(r"ou", clean, re.IGNORECASE):
                conjunto_id += 1
                continue
            m2 = re.match(r"^(\d{2})\s+(.+)$", clean)
            if m2:
                provas.append({
                    "conjunto_id": conjunto_id,
                    "code":        m2.group(1),
                    "name":        m2.group(2).strip(),
                })
    detail["provas"] = provas

    # ── Classificações mínimas ────────────────────────────────────────────────
    cmin = _sec_text(sections, "Classif")
    m = re.search(r"candidatura:\s*(\d+(?:[.,]\d)?)\s*pontos?", cmin)
    if m:
        detail["nota_minima_candidatura"] = float(m.group(1).replace(",", "."))
    m = re.search(r"[Pp]rovas? de ingresso:\s*(\d+(?:[.,]\d)?)\s*pontos?", cmin)
    if m:
        detail["nota_minima_prova"] = float(m.group(1).replace(",", "."))

    # ── Fórmula ───────────────────────────────────────────────────────────────
    formula = _sec_text(sections, "rmula")
    m = re.search(r"secund[aá]rio:\s*(\d+)\s*%", formula)
    if m:
        detail["peso_secundario"] = int(m.group(1)) / 100
    m = re.search(r"[Pp]rovas? de ingresso:\s*(\d+)\s*%", formula)
    if m:
        detail["peso_exame"] = int(m.group(1)) / 100

    # ── Historical grades (0-200 on page) ────────────────────────────────────
    stats_nodes = _sec_nodes(sections, "Dados Estat")
    if stats_nodes:
        stats_soup = BeautifulSoup("".join(str(n) for n in stats_nodes), "lxml")
        table = stats_soup.find("table")
        if table:
            trows = table.find_all("tr")
            if len(trows) >= 3:
                year_cells = [
                    td.get_text(strip=True)
                    for td in trows[0].find_all(["td", "th"])
                ]
                col_map: dict[int, tuple[int, int]] = {}
                col_idx = 1
                for yc in year_cells[1:]:
                    if re.match(r"\d{4}", yc):
                        yr = int(yc)
                        for fase in (1, 2):
                            col_map[col_idx] = (yr, fase)
                            col_idx += 1
                for data_row in trows[2:]:
                    cells = [
                        td.get_text(strip=True)
                        for td in data_row.find_all(["td", "th"])
                    ]
                    if not cells:
                        continue
                    if "ltimo Colocado" in cells[0] or "ltimo colocado" in cells[0]:
                        for ci, (yr, fase) in col_map.items():
                            if ci < len(cells) and cells[ci]:
                                v = _to_float(cells[ci])
                                if v is not None:
                                    detail[f"nota_{yr}_f{fase}"] = v  # 0-200 scale
                        break

    return detail

# ── Build output rows ─────────────────────────────────────────────────────────

def build_rows(
    pares: list[dict],
    notas_2025: dict,
    notas_2024: dict,
) -> tuple[list[dict], list[dict], list[dict]]:
    main_rows: list[dict]    = []
    provas_rows: list[dict]  = []
    hist_rows: list[dict]    = []

    for i, row in enumerate(pares):
        if i % 100 == 0:
            log.info("  %d / %d", i, len(pares))

        nome      = _s(row.get("CURSO"))
        inst_nome = _s(row.get("IES UO"))
        cod_uo    = _s(row.get("COD UO"))
        cod_ies   = _s(row.get("COD IES"))
        cod_curso = _s(row.get("COD CURSO"))
        if not nome or not cod_uo or not cod_curso:
            continue

        # Scrape (uses cache if available, fetches otherwise)
        detail = scrape_detail(cod_curso, cod_uo)

        subsistema = _s(row.get("SUBSISTEMA"))
        tipo = detail.get("tipo") or ("privada" if subsistema.lower() == "privado" else "publica")

        cnaef_code = _s(row.get("COD CNAEF"))
        area_app   = cnaef_to_area(cnaef_code)

        vagas_2026 = _i(row.get("REGIME GERAL DE ACESSO"))

        nota_2025_xl = notas_2025.get((cod_uo, cod_curso))
        nota_2024_xl = notas_2024.get((cod_uo, cod_curso))

        # Nota último colocado: HTML 2025 f1 → Excel 2025 → HTML 2024 f1 → Excel 2024
        nota_uc = (
            detail.get("nota_2025_f1")
            or nota_2025_xl
            or detail.get("nota_2024_f1")
            or nota_2024_xl
        )

        # Nota 2ª fase: most recent available
        nota_f2 = detail.get("nota_2025_f2") or detail.get("nota_2024_f2")

        # Provas string for readability
        from collections import defaultdict as _dd
        conj: dict = _dd(list)
        for p in detail.get("provas", []):
            conj[p["conjunto_id"]].append(f"{p['code']} {p['name']}")
        provas_str = "  OU  ".join(" + ".join(v) for _, v in sorted(conj.items()))

        main_rows.append({
            "Cód. UO":               cod_uo,
            "Cód. IES":              cod_ies,
            "Cód. Curso":            cod_curso,
            "Instituição":           inst_nome,
            "Curso":                 nome,
            "Grau":                  _s(row.get("TIPO CURSO")),
            "Tipo":                  tipo,
            "Área App":              area_app,
            "Área CNAEF (código)":   cnaef_code,
            "Subsistema":            subsistema,
            "Distrito":              detail.get("distrito", ""),
            # Vagas
            "Vagas 2026":            vagas_2026,
            # Notas último colocado (0-200 scale)
            "Nota 2023 - 1ª Fase":   detail.get("nota_2023_f1", ""),
            "Nota 2023 - 2ª Fase":   detail.get("nota_2023_f2", ""),
            "Nota 2024 - 1ª Fase":   detail.get("nota_2024_f1") or nota_2024_xl or "",
            "Nota 2024 - 2ª Fase":   detail.get("nota_2024_f2", ""),
            "Nota 2025 - 1ª Fase":   detail.get("nota_2025_f1") or nota_2025_xl or "",
            "Nota 2025 - 2ª Fase":   detail.get("nota_2025_f2", ""),
            # Nota principal (for import)
            "Nota Último Colocado":  nota_uc or "",
            "Nota Último Col. 2ªF":  nota_f2 or "",
            # Rules
            "Nota Mín. Candidatura": detail.get("nota_minima_candidatura", ""),
            "Nota Mín. Prova":       detail.get("nota_minima_prova", ""),
            "Peso Secundário":       detail.get("peso_secundario", ""),
            "Peso Exame":            detail.get("peso_exame", ""),
            # Provas
            "Provas de Ingresso":    provas_str,
            "N.º Conjuntos":         len(conj) if conj else "",
        })

        for p in detail.get("provas", []):
            provas_rows.append({
                "Cód. UO":    cod_uo,
                "Cód. Curso": cod_curso,
                "Instituição": inst_nome,
                "Curso":      nome,
                "Conjunto":   p["conjunto_id"],
                "Cód. Prova": p["code"],
                "Prova":      p["name"],
            })

        for yr in (2023, 2024, 2025):
            for fase in (1, 2):
                key = f"nota_{yr}_f{fase}"
                v = detail.get(key)
                if v is None and yr == 2025 and fase == 1:
                    v = nota_2025_xl
                if v is None and yr == 2024 and fase == 1:
                    v = nota_2024_xl
                if v is not None:
                    hist_rows.append({
                        "Cód. UO":    cod_uo,
                        "Cód. Curso": cod_curso,
                        "Instituição": inst_nome,
                        "Curso":      nome,
                        "Ano":        yr,
                        "Fase":       fase,
                        "Nota":       round(v, 1),
                    })

    return main_rows, provas_rows, hist_rows

# ── Excel styling ─────────────────────────────────────────────────────────────

def _style(ws):
    hfill = PatternFill("solid", fgColor="1A3A4A")
    afill = PatternFill("solid", fgColor="EDF4F7")
    for cell in ws[1]:
        cell.font      = Font(bold=True, color="FFFFFF", size=10)
        cell.fill      = hfill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for i, row in enumerate(ws.iter_rows(min_row=2), 2):
        fill = afill if i % 2 == 0 else None
        for cell in row:
            cell.alignment = Alignment(horizontal="left", vertical="center")
            if fill:
                cell.fill = fill
    for col in ws.columns:
        w = max((len(str(c.value or "")) for c in col), default=8)
        ws.column_dimensions[col[0].column_letter].width = min(w + 2, 55)
    ws.freeze_panes = "A2"
    ws.row_dimensions[1].height = 28

# ── Main ──────────────────────────────────────────────────────────────────────

def main():
    log.info("=== Step 1: Loading Excel sources ===")
    pares      = load_pares()
    notas_2025 = load_notas_2025()
    notas_2024 = load_notas_2024()

    log.info("=== Step 2: Scraping %d detail pages (uses cache when available) ===", len(pares))
    main_rows, provas_rows, hist_rows = build_rows(pares, notas_2025, notas_2024)

    log.info("=== Step 3: Writing Excel ===")
    import pandas as pd
    with pd.ExcelWriter(str(OUTPUT), engine="openpyxl") as w:
        pd.DataFrame(main_rows).to_excel(w,   index=False, sheet_name="Cursos")
        pd.DataFrame(provas_rows).to_excel(w, index=False, sheet_name="Provas (detalhe)")
        pd.DataFrame(hist_rows).to_excel(w,   index=False, sheet_name="Histórico")
        _style(w.sheets["Cursos"])
        _style(w.sheets["Provas (detalhe)"])
        _style(w.sheets["Histórico"])

    # Quality report
    total       = len(main_rows)
    with_nota25 = sum(1 for r in main_rows if r["Nota 2025 - 1ª Fase"] != "")
    with_nota24 = sum(1 for r in main_rows if r["Nota 2024 - 1ª Fase"] != "")
    with_provas = sum(1 for r in main_rows if r["Provas de Ingresso"])
    with_peso   = sum(1 for r in main_rows if r["Peso Secundário"] != "")
    with_dist   = sum(1 for r in main_rows if r["Distrito"])
    with_f2     = sum(1 for r in main_rows if r["Nota 2025 - 2ª Fase"] != "")

    log.info("=== Quality report ===")
    log.info("  Total courses:      %d", total)
    log.info("  Nota 2025 1ª fase:  %d / %d  (%.0f%%)", with_nota25, total, 100*with_nota25/total)
    log.info("  Nota 2025 2ª fase:  %d / %d  (%.0f%%)", with_f2,     total, 100*with_f2/total)
    log.info("  Nota 2024 1ª fase:  %d / %d  (%.0f%%)", with_nota24, total, 100*with_nota24/total)
    log.info("  Provas scraped:     %d / %d  (%.0f%%)", with_provas, total, 100*with_provas/total)
    log.info("  Peso scraped:       %d / %d  (%.0f%%)", with_peso,   total, 100*with_peso/total)
    log.info("  Distrito scraped:   %d / %d  (%.0f%%)", with_dist,   total, 100*with_dist/total)
    log.info("  Histórico rows:     %d", len(hist_rows))
    log.info("Saved → %s", OUTPUT)
    log.info("=== Done! ===")


if __name__ == "__main__":
    main()
